import sys
import unittest
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook


BACKEND_DIR = Path(__file__).resolve().parents[1] / "backend"
sys.path.insert(0, str(BACKEND_DIR))

from template_exporter import FLAT_SHEET_NAME, MAIN_SHEET_NAME, build_workbook  # noqa: E402


class TemplateExporterTests(unittest.TestCase):
    def test_exporter_writes_sheet2_and_fills_template(self):
        rows = [
            {
                "客户需求": "电子协议转订单",
                "一级模块": "协议识别",
                "二级模块": "内容识别",
                "三级模块": "电子协议内容识别",
                "功能用户": "发送者：客户经理 接受者：灵犀助手",
                "触发事件": "上传盖章协议后",
                "功能过程": "接收识别请求",
                "子过程描述": "捕获电子协议内容识别请求",
                "数据移动类型": "E",
                "数据组": "识别请求",
                "数据属性": "协议ID、文件数据、用户ID",
                "复用度": "新增",
                "CFP": 1,
            },
            {
                "客户需求": "电子协议转订单",
                "一级模块": "协议识别",
                "二级模块": "内容识别",
                "三级模块": "电子协议内容识别",
                "功能用户": "发送者：客户经理 接受者：灵犀助手",
                "触发事件": "上传盖章协议后",
                "功能过程": "读取识别规则",
                "子过程描述": "读取电子协议识别规则",
                "数据移动类型": "R",
                "数据组": "识别规则",
                "数据属性": "模型版本、字段映射、识别策略",
                "复用度": "复用",
                "CFP": 9,
            },
            {
                "客户需求": "电子协议转订单",
                "一级模块": "协议识别",
                "二级模块": "内容识别",
                "三级模块": "电子协议内容识别",
                "功能用户": "发送者：客户经理 接受者：灵犀助手",
                "触发事件": "上传盖章协议后",
                "功能过程": "复核历史识别结果",
                "子过程描述": "读取历史识别结果",
                "数据移动类型": "R",
                "数据组": "历史识别结果",
                "数据属性": "识别ID、协议ID、识别状态",
                "复用度": "利旧",
                "CFP": 9,
            },
        ]

        workbook = build_workbook(rows, "电子协议转订单")
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        saved = load_workbook(buffer, data_only=False)
        self.assertIn(FLAT_SHEET_NAME, saved.sheetnames)
        self.assertIn(MAIN_SHEET_NAME, saved.sheetnames)

        flat = saved[FLAT_SHEET_NAME]
        self.assertEqual(flat["A1"].value, "客户需求")
        self.assertEqual(flat["M2"].value, 1)

        main = saved[MAIN_SHEET_NAME]
        self.assertEqual(main["A5"].value, "电子协议转订单")
        self.assertEqual(main["G6"].value, "读取识别规则")
        self.assertEqual(main["M5"].value, 1)
        self.assertEqual(main["M6"].value, 0.33)
        self.assertEqual(main["M7"].value, 0)
        self.assertEqual(main["N5"].value, "=SUM(M5:M7)")
        self.assertEqual(main.max_row, 7)
        self.assertIn("A5:A7", [str(item) for item in main.merged_cells.ranges])
        self.assertIn("B5:B7", [str(item) for item in main.merged_cells.ranges])
        self.assertIn("C5:C7", [str(item) for item in main.merged_cells.ranges])
        self.assertIn("D5:D7", [str(item) for item in main.merged_cells.ranges])


if __name__ == "__main__":
    unittest.main()
