from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


BASE_DIR = Path(__file__).resolve().parents[1]
TEMPLATE_PATH = BASE_DIR / "灵犀助手--cosmic评估过程表 - 副本.xlsx"
MAIN_SHEET_NAME = "灵犀助手-cosmic "
FLAT_SHEET_NAME = "Sheet2"

EXPORT_HEADERS = [
    "客户需求",
    "一级模块",
    "二级模块",
    "三级模块",
    "功能用户",
    "触发事件",
    "功能过程",
    "子过程描述",
    "数据移动类型",
    "数据组",
    "数据属性",
    "复用度",
    "CFP",
]

MAIN_HEADERS = EXPORT_HEADERS + ["ΣCFP"]
MERGE_COLUMNS = [1, 2, 3, 4, 7]
PLAIN_STYLE_COLUMNS = [11]


def build_workbook(rows, project_name="COSMIC拆分结果"):
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"找不到模板文件：{TEMPLATE_PATH}")

    workbook = load_workbook(TEMPLATE_PATH)

    if FLAT_SHEET_NAME in workbook.sheetnames:
      del workbook[FLAT_SHEET_NAME]
    if "Sheet1" in workbook.sheetnames:
      del workbook["Sheet1"]

    flat_sheet = workbook.create_sheet(FLAT_SHEET_NAME)
    main_sheet = workbook[MAIN_SHEET_NAME]

    normalized_rows = [normalize_row(row, project_name) for row in rows]
    write_flat_sheet(flat_sheet, normalized_rows)
    write_main_sheet(main_sheet, normalized_rows)

    return workbook


def normalize_row(row, project_name):
    reuse = normalize_reuse(row.get("复用度"))
    cfp = reuse_to_cfp(reuse)

    return {
        "客户需求": str(row.get("客户需求") or project_name or "").strip(),
        "一级模块": str(row.get("一级模块") or "").strip(),
        "二级模块": str(row.get("二级模块") or "").strip(),
        "三级模块": str(row.get("三级模块") or "").strip(),
        "功能用户": str(row.get("功能用户") or "发送者：用户 接受者：灵犀助手").strip(),
        "触发事件": str(row.get("触发事件") or "").strip(),
        "功能过程": str(row.get("功能过程") or "").strip(),
        "子过程描述": str(row.get("子过程描述") or "").strip(),
        "数据移动类型": str(row.get("数据移动类型") or "").strip(),
        "数据组": str(row.get("数据组") or "").strip(),
        "数据属性": str(row.get("数据属性") or "").strip(),
        "复用度": reuse,
        "CFP": cfp,
    }


def reuse_to_cfp(reuse):
    text = normalize_reuse(reuse)
    if text == "复用":
        return 0.33
    if text == "利旧":
        return 0
    return 1


def normalize_reuse(reuse):
    text = str(reuse or "").strip()
    if "复用" in text:
        return "复用"
    if "利旧" in text:
        return "利旧"
    if "新增" in text:
        return "新增"
    return "新增"


def write_flat_sheet(sheet, rows):
    sheet.title = FLAT_SHEET_NAME
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = True

    header_row = 1
    for col_idx, header in enumerate(EXPORT_HEADERS, start=1):
        cell = sheet.cell(header_row, col_idx)
        cell.value = header
        style_flat_header(cell)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(EXPORT_HEADERS, start=1):
            cell = sheet.cell(row_idx, col_idx)
            cell.value = row[header]
            if header == "CFP":
                cell.number_format = "0.##"
            style_flat_cell(cell)

    widths = [14, 16, 18, 16, 30, 24, 18, 26, 10, 16, 32, 12, 10]
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + idx)].width = width


def write_main_sheet(sheet, rows):
    sheet.sheet_view.showGridLines = False
    _clear_main_data(sheet)
    _copy_template_row_style(sheet, source_row=5)

    data_start_row = 5
    data_end_row = data_start_row + len(rows) - 1

    for row_offset, row in enumerate(rows, start=data_start_row):
        _copy_template_row_style(sheet, source_row=5, target_row=row_offset)
        sheet.cell(row_offset, 1).value = row["客户需求"]
        sheet.cell(row_offset, 2).value = row["一级模块"]
        sheet.cell(row_offset, 3).value = row["二级模块"]
        sheet.cell(row_offset, 4).value = row["三级模块"]
        sheet.cell(row_offset, 5).value = row["功能用户"]
        sheet.cell(row_offset, 6).value = row["触发事件"]
        sheet.cell(row_offset, 7).value = row["功能过程"]
        sheet.cell(row_offset, 8).value = row["子过程描述"]
        sheet.cell(row_offset, 9).value = row["数据移动类型"]
        sheet.cell(row_offset, 10).value = row["数据组"]
        sheet.cell(row_offset, 11).value = row["数据属性"]
        sheet.cell(row_offset, 12).value = row["复用度"]
        sheet.cell(row_offset, 13).value = row["CFP"]
        sheet.cell(row_offset, 13).number_format = "0.##"
        sheet.cell(row_offset, 14).value = None

    if rows:
        sheet.cell(data_start_row, 14).value = f"=SUM(M{data_start_row}:M{data_end_row})"
        sheet.cell(data_start_row, 14).number_format = "0.##"

    if rows:
        _merge_contiguous(sheet, rows, data_start_row, data_end_row)

    _apply_main_style(sheet, data_start_row, data_end_row)
    _trim_unused_main_rows(sheet, data_end_row + 1)
    _set_main_dimensions(sheet)


def _clear_main_data(sheet):
    max_row = max(sheet.max_row, 5)
    for merged in list(sheet.merged_cells.ranges):
        if merged.min_row >= 5:
            sheet.unmerge_cells(str(merged))

    for row in sheet.iter_rows(min_row=5, max_row=max_row, min_col=1, max_col=14):
        for cell in row:
            cell.value = None


def _copy_template_row_style(sheet, source_row=5, target_row=None):
    target_row = source_row if target_row is None else target_row
    source = sheet[source_row]
    target = sheet[target_row]
    for source_cell, target_cell in zip(source[:14], target[:14]):
        target_cell._style = copy(source_cell._style)
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)

    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height


def _merge_contiguous(sheet, rows, start_row, end_row):
    for col_idx in MERGE_COLUMNS:
        key = _column_key(col_idx)
        run_start = start_row
        while run_start <= end_row:
            value = rows[run_start - start_row][key]
            run_end = run_start
            while run_end < end_row and rows[run_end - start_row + 1][key] == value:
                run_end += 1

            if value and run_end > run_start:
                sheet.merge_cells(start_row=run_start, start_column=col_idx, end_row=run_end, end_column=col_idx)
                top_left = sheet.cell(run_start, col_idx)
                top_left.alignment = copy(top_left.alignment)

            run_start = run_end + 1


def _apply_main_style(sheet, start_row, end_row):
    for row_idx in range(start_row, end_row + 1):
        for col_idx in range(1, 15):
            cell = sheet.cell(row_idx, col_idx)
            cell.alignment = Alignment(vertical="center", horizontal=cell.alignment.horizontal or "left", wrap_text=True)
            if col_idx in PLAIN_STYLE_COLUMNS:
                _clear_highlight(cell)

    sheet.freeze_panes = "A5"


def _trim_unused_main_rows(sheet, first_unused_row):
    if first_unused_row > sheet.max_row:
        return

    for merged in list(sheet.merged_cells.ranges):
        if merged.min_row >= first_unused_row:
            sheet.unmerge_cells(str(merged))

    sheet.delete_rows(first_unused_row, sheet.max_row - first_unused_row + 1)


def _set_main_dimensions(sheet):
    widths = {
        "A": 12.5,
        "B": 16.5,
        "C": 18.1640625,
        "D": 16.5,
        "E": 33.33203125,
        "F": 26.83203125,
        "G": 19.1640625,
        "H": 27.33203125,
        "I": 11.1640625,
        "J": 15.1640625,
        "K": 32.1640625,
        "L": 14.33203125,
        "M": 14.1640625,
        "N": 31.6640625,
    }
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width


def style_flat_header(cell):
    cell.fill = PatternFill("solid", fgColor="E8EEF7")
    cell.font = Font(bold=True, color="1F2937")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_flat_cell(cell):
    cell.alignment = Alignment(vertical="center", wrap_text=True)


def _column_key(index):
    mapping = {
        1: "客户需求",
        2: "一级模块",
        3: "二级模块",
        4: "三级模块",
        7: "功能过程",
    }
    return mapping[index]


def _clear_highlight(cell):
    base_font = copy(cell.font)
    cell.fill = PatternFill(fill_type=None)
    cell.font = Font(
        name=base_font.name,
        sz=base_font.sz,
        b=base_font.b,
        i=base_font.i,
        charset=base_font.charset,
        u=base_font.u,
        strike=base_font.strike,
        color="000000",
        vertAlign=base_font.vertAlign,
        family=base_font.family,
        scheme=base_font.scheme,
        outline=base_font.outline,
        shadow=base_font.shadow,
        condense=base_font.condense,
        extend=base_font.extend,
    )
